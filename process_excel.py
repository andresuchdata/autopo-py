import os
from pathlib import Path

import numpy as np
import pandas as pd

# Set base directory
BASE_DIR = Path(__file__).parent
XLSX_DIR = BASE_DIR / 'data' / 'rawpo' / 'xlsx'

OUTPUT_CSV_DIR = BASE_DIR / 'data' / 'rawpo' / 'input_csv'
OUTPUT_CSV_DIR.mkdir(parents=True, exist_ok=True)

NUMERIC_COLUMNS = [
    'HPP', 'Harga', 'Ranking', 'Grade', 'Terjual', 'Stok', 'Lost Days',
    'Velocity Capped', 'Daily Sales', 'Lead Time', 'Max. Daily Sales',
    'Max. Lead Time', 'Min. Order', 'Safety Stok', 'ROP', '3W Cover',
    'Sedang PO', 'Suggested', 'Amount', 'Promo Factor', 'Delay Factor',
    'Stock Cover', 'Days to Backup', 'Qty to Backup'
]

NA_VALUES = {
    'NAN', 'NA', '#N/A', 'NULL', 'NONE', '', '?', '-', 'INF', '-INF',
    '+INF', 'INFINITY', '-INFINITY', '1.#INF', '-1.#INF', '1.#QNAN'
}


def _patch_openpyxl_number_casting():
    """Ensure openpyxl won't crash when encountering NAN/INF in numeric cells."""
    try:
        from openpyxl.worksheet import _reader

        original_cast = _reader._cast_number

        def _safe_cast_number(value):  # pragma: no cover - monkey patch
            if isinstance(value, str):
                if value.strip().upper() in NA_VALUES:
                    return 0
            try:
                return original_cast(value)
            except (ValueError, TypeError):
                return 0 if value in (None, '') else value

        _reader._cast_number = _safe_cast_number
    except Exception:
        # If patch fails we continue; runtime reader will still attempt default behaviour
        pass


_patch_openpyxl_number_casting()

def clean_and_convert_v1(df):
    # Define numeric columns (case-insensitive)
    # Convert all columns to string first to avoid type issues
    df = df.astype(str)
    
    # Replace various NA/INF representations
    na_values = list(NA_VALUES)

    # Clean all columns
    for col in df.columns:
        # Check if column should be numeric (case-insensitive match)
        is_numeric = any(nc.lower() in col.lower() for nc in NUMERIC_COLUMNS)
        
        if is_numeric:
            # For numeric columns, replace NA/INF with 0
            df[col] = pd.to_numeric(
                df[col].replace(na_values, np.nan, regex=True),
                errors='coerce'
            ).fillna(0)
        else:
            # For non-numeric columns, replace NA/INF with empty string
            df[col] = df[col].replace(na_values, '', regex=True)
    
    return df

def clean_and_convert(df):
    """Clean and convert DataFrame columns to appropriate types."""
    if df is None or df.empty:
        return df

    # Make a copy to avoid SettingWithCopyWarning
    df = df.copy()
    
    # Convert all columns to string first to handle NaN/None consistently
    for col in df.columns:
        df[col] = df[col].astype(str)
    
    # Define NA values that should be treated as empty/missing
    na_values = list(NA_VALUES)
    
    # Process each column
    for col in df.columns:
        # Replace NA values with empty string first (treating them as literals, not regex)
        df[col] = df[col].replace(na_values, '', regex=False)
        
        # Skip empty columns
        if df[col].empty:
            continue

        # Convert numeric columns
        if col in NUMERIC_COLUMNS:
            # Convert to numeric, coercing errors to NaN, then fill with 0
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            # For non-numeric columns, ensure they're strings and strip whitespace
            df[col] = df[col].astype(str).str.strip()
            # Replace empty strings with NaN and then fill with empty string
            df[col] = df[col].replace('', np.nan).fillna('')

    return df

def read_excel_file(file_path):
    """
    Read an Excel file with robust error handling for problematic values.
    """
    try:
        print(f"\nProcessing {file_path.name}...")
        
        # First, read the file with openpyxl directly to handle the data more carefully
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(
            filename=file_path,
            read_only=True,    # Read-only mode is faster and uses less memory
            data_only=True,    # Get the stored value instead of the formula
            keep_links=False   # Don't load external links
        )
        
        # Get the first sheet
        ws = wb.active
        
        # Get headers from the first row
        headers = []
        for idx, cell in enumerate(next(ws.iter_rows(values_only=True))):
            header = str(cell).strip() if cell not in (None, '') else f"Column_{idx + 1}"
            headers.append(header)
        
        # Initialize data rows
        data = []
        
        # Process each row
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                    continue

                cell_str = str(cell).strip()
                if cell_str.upper() in NA_VALUES:
                    row_data.append('')
                else:
                    row_data.append(cell_str)
            
            # Only add row if it has data
            if any(cell != '' for cell in row_data):
                data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Normalize column data types
        df = clean_and_convert(df)
        
        print(f"✅ Successfully processed {file_path.name} with {len(df)} rows")
        return df
        
    except Exception as e:
        print(f"❌ Error processing {file_path.name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def process_excel_files():
    """Process all Excel files in the xlsx directory."""
    # Get all xlsx files
    xlsx_files = list(XLSX_DIR.glob('*.xlsx'))
    
    if not xlsx_files:
        print(f"No Excel files found in {XLSX_DIR}")
        return
    
    print(f"Found {len(xlsx_files)} Excel files to process...")
    
    # Process each file
    for file_path in xlsx_files:
        print(f"\nProcessing {file_path.name}...")
        
        # Read the Excel file
        df = read_excel_file(file_path)
        
        if df is not None and not df.empty:
            print(f"Successfully read {file_path.name} with {len(df)} rows"
                  f" and {len(df.columns)} columns")
            print("\nFirst few rows:")
            print(df.head())
            
            # Here you can add your processing logic
            # For example, you might want to clean the data, transform it, etc.
            
        else:
            print(f"Failed to process {file_path.name}")
    
    print("\nProcessing complete!")

if __name__ == "__main__":
    # Process all Excel files in the directory
    xlsx_files = list(XLSX_DIR.glob('*.xlsx'))
    
    if not xlsx_files:
        print(f"No Excel files found in {XLSX_DIR}")
    else:
        for file_path in xlsx_files:
            print(f"\nProcessing {file_path.name}...")
            df = read_excel_file(file_path)

            
            if df is not None:
                print(f"✅ Successfully read {file_path.name} with {len(df)} rows")
                # print(f"Columns: {list(df.columns)}")

                # Save to CSV
                output_path = OUTPUT_CSV_DIR / f"{file_path.stem}.csv"
                df.to_csv(output_path, index=False)
                print(f"✅ Saved to {output_path}")
            else:
                print(f"❌ Failed to read {file_path.name}")
