
import numpy as np
from dotenv import load_dotenv
import pandas as pd
import os
from pathlib import Path
from locale import atof
from openpyxl.styles import numbers
from datetime import datetime
import openpyxl.utils
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

# Load env if available
load_dotenv()

BASE_DIR = Path('/Users/andresuchitra/dev/missglam/autopo/notebook')

def _normalize_col(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    norm_to_actual = {_normalize_col(c): c for c in df.columns}
    for cand in candidates:
        key = _normalize_col(cand)
        if key in norm_to_actual:
            return norm_to_actual[key]

    return None


def _default_compare_config() -> dict:
    return {
        "mode": "strict",  # strict | loose
        "number_locale": "auto",  # auto | id | us
        "abs_tol": 0.0,
        "rel_tol": 0.0,
        "treat_blank_as_zero": False,
        "treat_blank_as_null": True,
        "na_tokens": ["", "nan", "none", "null", "na", "#n/a"],
        # Merge-key normalization
        "normalize_keys": True,
        "normalize_sku": True,
        "normalize_toko": True,
        "toko_key_mode": "slug",  # slug | simple
    }


def _parse_number_one(x, locale: str = "auto"):
    if x is None:
        return None
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return None

    # Remove internal spaces often found in ID format (e.g. "1 000")
    s = s.replace(" ", "")

    def parse_id(ss: str):
        ss = ss.replace(".", "")
        ss = ss.replace(",", ".")
        return float(ss)

    def parse_us(ss: str):
        ss = ss.replace(",", "")
        return float(ss)

    try:
        if locale == "id":
            return parse_id(s)
        if locale == "us":
            return parse_us(s)

        has_comma = "," in s
        has_dot = "." in s

        if has_comma and has_dot:
            if s.rfind(",") > s.rfind("."):
                return parse_id(s)
            return parse_us(s)

        if has_comma and not has_dot:
            if s.count(",") > 1:
                return parse_us(s)
            # Single comma: likely ID decimal unless it looks like a US thousand separator
            # Heuristic: if strictly 3 digits after comma, could be US thousand. 
            # But in this dataset, "," is clearly used for decimals (0,34).
            # So we prefer ID parse if it looks like a decimal number.
            if re.fullmatch(r"-?\d+,\d+", s):
                return parse_id(s)
            return parse_us(s)

        if has_dot and not has_comma:
            if s.count(".") > 1:
                return parse_id(s)
            m = re.fullmatch(r"-?\d+\.(\d+)", s)
            if m and len(m.group(1)) == 3:
                return parse_id(s)
            return parse_us(s)

        return float(s)
    except Exception:
        return None


def _to_number_series(s: pd.Series, compare_config: dict | None = None) -> pd.Series:
    if s is None:
        return s
    cfg = _default_compare_config() if compare_config is None else {**_default_compare_config(), **compare_config}

    s2 = s.astype(str).str.strip()

    na_tokens = set([str(t).strip().lower() for t in (cfg.get("na_tokens") or [])])
    if cfg.get("treat_blank_as_null", True):
        # Map values present in na_tokens to None
        s2 = s2.map(lambda v: None if str(v).strip().lower() in na_tokens else v)

    if cfg.get("treat_blank_as_zero"):
        s2 = s2.fillna("0") # fillna handles None from previous step

    locale = cfg.get("number_locale", "auto")
    return s2.map(lambda v: _parse_number_one(v, locale=locale))


def _plain_sum_series(df: pd.DataFrame, column: str, compare_config: dict | None = None) -> float:
    if df is None or column not in df.columns:
        return 0.0

    # FIX: Use robust parsing instead of simple pd.to_numeric
    # series = pd.to_numeric(df[column], errors="coerce")
    series = _to_number_series(df[column], compare_config=compare_config)
    series = series.fillna(0)
    return float(series.sum())


def compare_series(a: pd.Series, b: pd.Series, merge_indicator: pd.Series | None, compare_config: dict | None = None) -> pd.Series:
    cfg = _default_compare_config() if compare_config is None else {**_default_compare_config(), **compare_config}

    abs_tol = float(cfg.get("abs_tol", 0.0) or 0.0)
    rel_tol = float(cfg.get("rel_tol", 0.0) or 0.0)

    if cfg.get("mode") == "loose" and abs_tol == 0.0 and rel_tol == 0.0:
        abs_tol = 1e-9

    aa = _to_number_series(a, compare_config=cfg)
    bb = _to_number_series(b, compare_config=cfg)

    both_null = aa.isna() & bb.isna()
    one_null = aa.isna() ^ bb.isna()

    diff = (aa - bb).abs()
    tol = abs_tol + rel_tol * bb.abs().fillna(0.0)
    same_numeric = (~both_null) & (~one_null) & (diff <= tol)

    same = both_null | same_numeric

    if merge_indicator is not None:
        same = same & merge_indicator.eq("both")

    return same

def _normalize_sku_key_series(s: pd.Series) -> pd.Series:
    s2 = s.astype(str).str.strip()
    s2 = s2.str.replace(r"\s+", "", regex=True)
    s2 = s2.replace({"nan": "", "None": "", "": ""})
    return s2

def _normalize_store_key_series_slug(s: pd.Series) -> pd.Series:
    # Similar to main.ipynb: make merges robust against punctuation/spacing differences
    s2 = s.astype(str).str.strip().str.lower()
    s2 = s2.str.replace(r"\s+", " ", regex=True)
    s2 = s2.str.replace(r"[^a-z0-9]+", "", regex=True)
    s2 = s2.replace({"nan": "", "none": "", "": ""})
    return s2

def _normalize_store_key_series_simple(s: pd.Series) -> pd.Series:
    s2 = s.astype(str).str.strip()
    s2 = s2.str.replace(r"\s+", " ", regex=True)
    s2 = s2.replace({"nan": "", "None": "", "": ""})
    return s2

def _apply_key_normalization(df: pd.DataFrame, compare_config: dict | None = None) -> pd.DataFrame:
    cfg = _default_compare_config() if compare_config is None else {**_default_compare_config(), **compare_config}
    if not cfg.get("normalize_keys", True):
        return df

    if cfg.get("normalize_sku", True) and "SKU" in df.columns:
        df["SKU"] = _normalize_sku_key_series(df["SKU"])

    if cfg.get("normalize_toko", True) and "Toko" in df.columns:
        mode = str(cfg.get("toko_key_mode", "slug") or "slug").lower()
        if mode == "simple":
            df["Toko"] = _normalize_store_key_series_simple(df["Toko"])
        else:
            df["Toko"] = _normalize_store_key_series_slug(df["Toko"])

    return df

def validate_po_fields(
    input_csv_path: str | Path,
    output_csv_path: str | Path,
    key_cols: tuple[str, str] = ("SKU", "Toko"),
    fields: list[str] | None = None,
    compare_config: dict | None = None,
):
    if fields is None:
        fields = [
            "Stok",
            "Daily Sales",
            "Max. Daily Sales",
            "Lead Time",
            "Max. Lead Time",
            "Sedang PO",
        ]

    # These are extra columns we want to ensure are passed through to the merged DF
    # for display purposes even if not strictly compared in the "fields" loop.
    extra_display_fields = ["emergency_po_cost", "final_updated_regular_po_cost"]

    input_csv_path = Path(input_csv_path)
    output_csv_path = Path(output_csv_path)

    df_in = _read_input_csv_any_encoding(input_csv_path, sep=",", dtype=str, keep_default_na=False)
    df_out = pd.read_csv(output_csv_path, sep=";", dtype=str, keep_default_na=False)

    in_sku = _find_col(df_in, [key_cols[0]])
    in_toko = _find_col(df_in, [key_cols[1]])
    out_sku = _find_col(df_out, [key_cols[0]])
    out_toko = _find_col(df_out, [key_cols[1]])

    missing_keys = [k for k, v in [("input.SKU", in_sku), ("input.Toko", in_toko), ("output.SKU", out_sku), ("output.Toko", out_toko)] if v is None]
    if missing_keys:
        raise ValueError(
            f"Missing key columns: {missing_keys}. Available input cols={list(df_in.columns)[:20]}..., output cols={list(df_out.columns)[:20]}..."
        )

    df_in = df_in.rename(columns={in_sku: "SKU", in_toko: "Toko"})
    df_out = df_out.rename(columns={out_sku: "SKU", out_toko: "Toko"})

    # Preserve originals for reporting
    df_in["SKU__orig"] = df_in["SKU"]
    df_in["Toko__orig"] = df_in["Toko"]
    df_out["SKU__orig"] = df_out["SKU"]
    df_out["Toko__orig"] = df_out["Toko"]

    # Normalize merge keys so SKU+Toko truly matches between files
    df_in = _apply_key_normalization(df_in, compare_config=compare_config)
    df_out = _apply_key_normalization(df_out, compare_config=compare_config)

    out_stock_col = _find_col(df_out, ["Stock", "Stok"])
    if out_stock_col is None:
        raise ValueError("Output is missing 'Stock'/'Stok' column")

    col_map = {
        "Stok": ("Stok", out_stock_col),
        "Daily Sales": ("Daily Sales", "Daily Sales"),
        "Max. Daily Sales": ("Max. Daily Sales", "Max. Daily Sales"),
        "Lead Time": ("Lead Time", "Lead Time"),
        "Max. Lead Time": ("Max. Lead Time", "Max. Lead Time"),
        "Sedang PO": ("Sedang PO", "Sedang PO"),
    }

    missing = []
    for f in fields:
        if f not in col_map:
            missing.append(f"Unknown field '{f}'")
            continue
        in_col, out_col = col_map[f]
        if _find_col(df_in, [in_col]) is None:
            missing.append(f"input.{in_col}")
        if _find_col(df_out, [out_col]) is None:
            missing.append(f"output.{out_col}")
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    rename_in = {}
    rename_out = {}
    for f in fields:
        in_col, out_col = col_map[f]
        actual_in = _find_col(df_in, [in_col])
        actual_out = _find_col(df_out, [out_col])
        rename_in[actual_in] = f"in__{f}"
        rename_out[actual_out] = f"out__{f}"

    # Handle extra display fields
    for f in extra_display_fields:
        if f in df_out.columns:
            rename_out[f] = f"out_extra__{f}"
    
    df_in2 = df_in.rename(columns=rename_in)
    df_out2 = df_out.rename(columns=rename_out)

    merged = df_in2.merge(df_out2, on=["SKU", "Toko"], how="outer", indicator=True, suffixes=("__in", "__out"))

    # Restore best-effort original key strings for report readability
    if "SKU__orig__in" in merged.columns and "SKU__orig__out" in merged.columns:
        merged["SKU"] = merged["SKU__orig__in"].where(
            merged["SKU__orig__in"].astype(str).str.strip() != "", merged["SKU__orig__out"]
        )
    if "Toko__orig__in" in merged.columns and "Toko__orig__out" in merged.columns:
        merged["Toko"] = merged["Toko__orig__in"].where(
            merged["Toko__orig__in"].astype(str).str.strip() != "", merged["Toko__orig__out"]
        )

    mismatch_rows = []
    for f in fields:
        same = compare_series(
            merged.get(f"in__{f}"),
            merged.get(f"out__{f}"),
            merged.get("_merge"),
            compare_config=compare_config,
        )
        mask = ~same

        if mask.any():
            tmp = merged.loc[mask, ["SKU", "Toko", "_merge", f"in__{f}", f"out__{f}"]].copy()
            tmp.insert(2, "field", f)
            tmp = tmp.rename(columns={f"in__{f}": "input_value", f"out__{f}": "output_value"})
            mismatch_rows.append(tmp)

    mismatches = (
        pd.concat(mismatch_rows, ignore_index=True)
        if mismatch_rows
        else pd.DataFrame(columns=["SKU", "Toko", "field", "_merge", "input_value", "output_value"])
    )
    return merged, mismatches


def build_validation_flags(
    merged_df: pd.DataFrame,
    fields: list[str] | None = None,
    compare_config: dict | None = None,
) -> pd.DataFrame:
    if fields is None:
        fields = [
            "Stok",
            "Daily Sales",
            "Max. Daily Sales",
            "Lead Time",
            "Max. Lead Time",
            "Sedang PO",
        ]

    out = merged_df[["SKU", "Toko", "_merge"]].copy()

    for f in fields:
        in_col = f"in__{f}"
        out_col = f"out__{f}"

        out[f"input__{f}"] = merged_df.get(in_col)
        out[f"output__{f}"] = merged_df.get(out_col)

        same = compare_series(
            merged_df.get(in_col),
            merged_df.get(out_col),
            merged_df.get("_merge"),
            compare_config=compare_config,
        )

        out[f"same__{f}"] = same.map(lambda x: "Y" if bool(x) else "N")

    # Add extra columns for display
    # "emergency_po_cost" and "final_updated_regular_po_cost"
    extra_cols = ["emergency_po_cost", "final_updated_regular_po_cost"]
    for c in extra_cols:
        col_key = f"out_extra__{c}"
        if col_key in merged_df.columns:
             # Just copy the value as a single column.
             # We name it exactly as requested
             out[c] = merged_df[col_key]

    return out


def export_validation_xlsx(
    merged_df: pd.DataFrame,
    output_xlsx_path: str | Path,
    fields: list[str] | None = None,
    compare_config: dict | None = None,
    sheet_name: str = "validation",
) -> Path:
    output_xlsx_path = Path(output_xlsx_path)
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    flags_df = build_validation_flags(merged_df, fields=fields, compare_config=compare_config)

    with pd.ExcelWriter(output_xlsx_path, engine="openpyxl") as writer:
        flags_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column
        ws.auto_filter.ref = ws.cell(row=1, column=1).coordinate + ":" + ws.cell(row=max_row, column=max_col).coordinate

    return output_xlsx_path


def _read_input_csv_any_encoding(path: Path, /, *, sep: str = ",", **read_kwargs) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
    last_err = None
    for enc in encodings:
        try:
            return pd.read_csv(path, sep=sep, encoding=enc, **read_kwargs)
        except UnicodeDecodeError as err:
            last_err = err
    if last_err is not None:
        raise last_err

    return pd.read_csv(path, sep=sep, **read_kwargs)

def _list_input_files_for_date(date_str: str, input_base_dir: Path = BASE_DIR / "data/input") -> list[Path]:
    input_dir = input_base_dir / date_str
    if not input_dir.exists():
        raise FileNotFoundError(f"Input dir not found: {input_dir}")

    files = sorted([p for p in input_dir.iterdir() if p.is_file() and p.suffix.lower() == ".csv"])
    return files


def _resolve_output_file(
    input_file: Path,
    date_str: str,
    output_base_dir: Path = BASE_DIR / "output",
    output_subdir: str = "complete/csv",
    output_exts: list[str] | None = None,
) -> Path | None:
    # Expected output structure: output/YYYYMMDD/complete/csv/<same filename>
    output_dir = output_base_dir / date_str / output_subdir
    
    if not output_dir.exists():
        return None

    if output_exts is None:
        output_exts = [".csv"]

    candidate = output_dir / input_file.name
    if candidate.exists():
        return candidate

    for ext in output_exts:
        cand2 = output_dir / f"{input_file.stem}{ext}"
        if cand2.exists():
            return cand2

    return None


def _normalize_date_string(date_str: str) -> str:
    month_mapping = {
        "jan": "jan", "januari": "january",
        "feb": "feb", "februari": "february",
        "mar": "mar", "maret": "march",
        "apr": "apr", "april": "april",
        "mei": "may", "may": "may",
        "jun": "jun", "juni": "june",
        "jul": "jul", "juli": "july",
        "agu": "aug", "agustus": "august",
        "sep": "sep", "september": "september",
        "okt": "oct", "oktober": "october",
        "nov": "nov", "november": "november",
        "des": "dec", "desember": "december",
        "dec": "dec",
    }

    s = str(date_str).lower().strip()
    for indo, eng in month_mapping.items():
        s = re.sub(r"\b" + re.escape(indo) + r"\b", eng, s)
    return s


def _find_date_header_columns(df: pd.DataFrame, target_date_str: str) -> list[str]:
    normalized_target = _normalize_date_string(target_date_str)
    target_ts = pd.to_datetime(normalized_target, dayfirst=True, errors="coerce")
    if pd.isna(target_ts):
        return []

    matched = []
    for col in df.columns:
        col_str = str(col).strip()
        col_ts = pd.to_datetime(_normalize_date_string(col_str), dayfirst=True, errors="coerce")
        if pd.isna(col_ts):
            continue
        if col_ts.normalize() == target_ts.normalize():
            matched.append(col)
    return matched


def _build_metrics_df(in_file: Path, out_file: Path, compare_config: dict | None = None) -> pd.DataFrame:
    df_in = _read_input_csv_any_encoding(in_file, sep=",", dtype=str, keep_default_na=False)
    df_out = pd.read_csv(out_file, sep=";", dtype=str, keep_default_na=False)

    in_sku = _find_col(df_in, ["SKU"])
    in_toko = _find_col(df_in, ["Toko"])
    out_sku = _find_col(df_out, ["SKU"])
    out_toko = _find_col(df_out, ["Toko"])

    if in_sku is None or in_toko is None or out_sku is None or out_toko is None:
        return pd.DataFrame([
            {"metric": "error", "value": "Missing SKU/Toko key columns in input or output"}
        ])

    df_in = df_in.rename(columns={in_sku: "SKU", in_toko: "Toko"})
    df_out = df_out.rename(columns={out_sku: "SKU", out_toko: "Toko"})

    df_in = _apply_key_normalization(df_in, compare_config=compare_config)
    df_out = _apply_key_normalization(df_out, compare_config=compare_config)

    def _num_series(frame: pd.DataFrame, col: str) -> pd.Series:
        if col not in frame.columns:
            return pd.Series([np.nan] * len(frame))
        return _to_number_series(frame[col], compare_config=compare_config)
    
    # helper for sum
    def _sum_num(frame: pd.DataFrame, col: str) -> float:
        return float(_num_series(frame, col).fillna(0).sum())

    def _minmax_num(frame: pd.DataFrame, col: str):
        if col not in frame.columns:
            return None, None
        s = _num_series(frame, col).dropna()
        if s.empty:
            return None, None
        return float(s.min()), float(s.max())

    final_cost_min, final_cost_max = _minmax_num(df_out, "final_updated_regular_po_cost")
    final_qty_min, final_qty_max = _minmax_num(df_out, "final_updated_regular_po_qty")

    emergency_cost_min, emergency_cost_max = _minmax_num(df_out, "emergency_po_cost")
    emergency_qty_min, emergency_qty_max = _minmax_num(df_out, "emergency_po_qty")
    
    # sums
    sum_final_updated_po_cost = _sum_num(df_out, "final_updated_regular_po_cost")
    sum_emergency_po_cost = _sum_num(df_out, "emergency_po_cost")

    stock_col = "final_stock" if "final_stock" in df_out.columns else (_find_col(df_out, ["Stock", "Stok"]) or "")
    final_stock_min, final_stock_max = (None, None)
    if stock_col and stock_col in df_out.columns:
        final_stock_min, final_stock_max = _minmax_num(df_out, stock_col)

    qty_same_all = None
    qty_same_count = None
    qty_mismatch_count = None
    if "updated_regular_po_qty" in df_out.columns and "initial_qty_po" in df_out.columns:
        a = _num_series(df_out, "updated_regular_po_qty").fillna(0)
        b = _num_series(df_out, "initial_qty_po").fillna(0)
        same = a.eq(b)
        qty_same_all = bool(same.all())
        qty_same_count = int(same.sum())
        qty_mismatch_count = int((~same).sum())

    cover_max_abs_diff = None
    cover_mismatch_count = None
    cover_computed = False

    if stock_col and stock_col in df_out.columns and "current_days_stock_cover" in df_out.columns:
        in_daily_sales_col = _find_col(df_in, ["Daily Sales"])
        if in_daily_sales_col is not None:
            tmp_in = df_in[["SKU", "Toko", in_daily_sales_col]].copy()
            tmp_in = tmp_in.rename(columns={in_daily_sales_col: "Daily Sales__input"})
            merged = df_out.merge(tmp_in, on=["SKU", "Toko"], how="left")

            st = _num_series(merged, stock_col).fillna(0).to_numpy()
            ds_in = _num_series(merged, "Daily Sales__input").fillna(0).to_numpy()
            cover_in = np.where(ds_in > 0, st / ds_in, 0)

            cover_out = _num_series(merged, "current_days_stock_cover").fillna(0).to_numpy()
            diff = np.abs(cover_out - cover_in)
            cover_max_abs_diff = float(np.nanmax(diff)) if diff.size else None
            cover_mismatch_count = int(np.sum(diff > 1e-6)) if diff.size else None
            cover_computed = True

    rows = [
        ("file", in_file.name),
        ("final_regular_po_cost_min", final_cost_min),
        ("final_regular_po_cost_max", final_cost_max),
        ("sum_final_updated_po_cost", sum_final_updated_po_cost),
        ("final_updated_po_qty_min", final_qty_min),
        ("final_updated_po_qty_max", final_qty_max),
        ("emergency_po_cost_min", emergency_cost_min),
        ("emergency_po_cost_max", emergency_cost_max),
        ("sum_emergency_po_cost", sum_emergency_po_cost),
        ("emergency_po_qty_min", emergency_qty_min),
        ("emergency_po_qty_max", emergency_qty_max),
        ("final_stock_min", final_stock_min),
        ("final_stock_max", final_stock_max),
        ("updated_vs_initial_qty_all_same", qty_same_all),
        ("updated_vs_initial_qty_same_count", qty_same_count),
        ("updated_vs_initial_qty_mismatch_count", qty_mismatch_count),
        ("current_days_stock_cover_diff_computed", bool(cover_computed)),
        ("current_days_stock_cover_max_abs_diff", cover_max_abs_diff),
        ("current_days_stock_cover_mismatch_count", cover_mismatch_count),
    ]

    return pd.DataFrame(rows, columns=["metric", "value"])


def _resolve_top100_dir_simple() -> Path | None:
    env_path = os.environ.get("TOP100_GDRIVE_DIR")
    if env_path:
        p = Path(env_path).expanduser()
        if p.exists() and p.is_dir():
            return p

    local_top100 = BASE_DIR / "data/top_100_sku"
    if local_top100.exists() and local_top100.is_dir():
        return local_top100

    return None


def _get_store_name_from_filename(filename: str) -> str:
    name_parts = Path(filename).stem.split()

    if len(name_parts) >= 3 and name_parts[1].lower() == 'miss' and name_parts[2].lower() == 'glam':
        return ' '.join(name_parts[3:]).strip().upper()
    if len(name_parts) >= 2 and name_parts[0].lower() == 'miss' and name_parts[1].lower() == 'glam':
        return ' '.join(name_parts[2:]).strip().upper()
    if ' ' in filename:
        return ' '.join(name_parts[1:]).strip().upper()
    return name_parts[0].upper()


TOP100_SKU_CACHE = None


def _preload_all_top_100_sku(top_100_dir: Path, max_rows: int = 100) -> dict[str, pd.DataFrame]:
    store_names = []
    for path in sorted(top_100_dir.glob('*')):
        if not path.is_file():
            continue
        store_names.append(_get_store_name_from_filename(path.name))

    store_names = sorted(set([str(s).strip().upper() for s in store_names if str(s).strip()]))

    cache: dict[str, pd.DataFrame] = {}
    for store_name in store_names:
        df_store = _load_top_100_sku_for_store(store_name, top_100_dir=top_100_dir, max_rows=max_rows)
        if df_store is not None and not df_store.empty:
            cache[store_name] = df_store

    return cache


def _load_top_100_sku_for_store(location: str, top_100_dir: Path, max_rows: int = 100) -> pd.DataFrame | None:
    location_upper = str(location).strip().upper()

    global TOP100_SKU_CACHE
    if isinstance(TOP100_SKU_CACHE, dict) and location_upper in TOP100_SKU_CACHE:
        return TOP100_SKU_CACHE[location_upper]

    matching_files = []
    for path in sorted(top_100_dir.glob('*')):
        if not path.is_file():
            continue
        store_name = _get_store_name_from_filename(path.name)
        if store_name.upper() == location_upper:
            matching_files.append(path)

    if not matching_files:
        return None

    file_path = matching_files[0]

    suffix = file_path.suffix.lower()
    if suffix in ['.xlsx', '.xls']:
        raw_df = pd.read_excel(file_path, header=None, engine='openpyxl')
    else:
        read_ok = False
        raw_df = None
        for sep in [';', ',']:
            for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
                try:
                    raw_df = pd.read_csv(file_path, header=None, sep=sep, encoding=enc)
                    read_ok = True
                    break
                except Exception:
                    continue
            if read_ok:
                break

        if not read_ok or raw_df is None:
            return None

    if raw_df is None or raw_df.empty:
        return None

    header_row = None
    max_header_search = min(15, len(raw_df))

    for idx in range(max_header_search):
        row_values = raw_df.iloc[idx].astype(str).str.strip()
        lowered = [v.lower() for v in row_values]
        if "sku" in lowered:
            header_row = idx
            break

    if header_row is None:
        for idx in range(max_header_search):
            if raw_df.iloc[idx].notna().any():
                header_row = idx
                break

    if header_row is None:
        header_row = 0

    header = raw_df.iloc[header_row].astype(str).str.strip().tolist()
    data = raw_df.iloc[header_row + 1 :].reset_index(drop=True)
    data.columns = header
    df = data

    df.columns = df.columns.astype(str).str.strip()

    sku_col = None
    for c in df.columns:
        if str(c).strip().lower() == "sku":
            sku_col = c
            break

    if sku_col is None:
        return None

    df[sku_col] = df[sku_col].astype(str).str.strip()
    if sku_col != "SKU":
        df["SKU"] = df[sku_col]

    df = df.dropna(how='all')

    if max_rows is not None and max_rows > 0:
        df = df.head(max_rows)

    return df

def _ensure_top100_cache_loaded() -> None:
    global TOP100_SKU_CACHE
    if isinstance(TOP100_SKU_CACHE, dict):
        return

    top100_dir = _resolve_top100_dir_simple()
    if top100_dir is None:
        TOP100_SKU_CACHE = {}
        return

    TOP100_SKU_CACHE = _preload_all_top_100_sku(top_100_dir=top100_dir, max_rows=100)

def _resolve_top100_dir() -> Path | None:
    # backward-compatible alias
    return _resolve_top100_dir_simple()


def _read_top100_store_file(top100_dir: Path, store_stem: str) -> Path | None:
    patterns = [
        f"{store_stem}.xlsx",
        f"{store_stem}.xls",
        f"{store_stem}.csv",
    ]
    for pat in patterns:
        p = top100_dir / pat
        if p.exists() and p.is_file():
            return p

    matches = []
    for ext in (".xlsx", ".xls", ".csv"):
        matches.extend(list(top100_dir.glob(f"*{store_stem}*{ext}")))

    matches = [m for m in matches if m.is_file()]
    matches = sorted(matches)
    return matches[0] if matches else None


def _truthy_flag_series(s: pd.Series) -> pd.Series:
    s2 = s.astype(str).str.strip().str.lower()
    return s2.isin(["1", "true", "y", "yes", "t"])


def _build_top100_comparison_df(
    in_file: Path,
    out_file: Path,
    date_str: str,
    compare_config: dict | None = None,
) -> pd.DataFrame:
    top100_dir = _resolve_top100_dir()
    if top100_dir is None:
        return pd.DataFrame([
            {"error": "Top100 dir not found. Set env TOP100_GDRIVE_DIR or ensure BASE_DIR/data/top_100_sku exists."}
        ])

    store_stem = out_file.stem
    top_file = _read_top100_store_file(top100_dir, store_stem)
    if top_file is None:
        return pd.DataFrame([
            {"error": f"Top100 store file not found in {str(top100_dir)} for stem '{store_stem}'"}
        ])

    df_out = pd.read_csv(out_file, sep=";", dtype=str, keep_default_na=False)

    out_sku = _find_col(df_out, ["SKU"])
    out_toko = _find_col(df_out, ["Toko"])
    if out_sku is None or out_toko is None:
        return pd.DataFrame([{ "error": "Output missing SKU/Toko" }])

    df_out = df_out.rename(columns={out_sku: "SKU", out_toko: "Toko"})
    df_out = _apply_key_normalization(df_out, compare_config=compare_config)

    # Filter ONLY top100 SKUs from output
    top100_flag_col = "is_top_100_sku"

    top100_mask = _truthy_flag_series(df_out[top100_flag_col])
    df_out = df_out.loc[top100_mask].copy()

    if df_out.empty:
        return pd.DataFrame([
            {"error": f"No output rows with {top100_flag_col}=true/1 for file {out_file.name}"}
        ])

    # Output columns
    final_stock_col = "final_stock" if "final_stock" in df_out.columns else (_find_col(df_out, ["Stock", "Stok"]) or None)
    stok_col = _find_col(df_out, ["Stok", "Stock"])  # for display + fallback comparison

    if final_stock_col is None:
        return pd.DataFrame([{ "error": "Output missing final_stock/Stock/Stok" }])

    df_out_stock = df_out[["SKU", "Toko", final_stock_col]].copy()
    df_out_stock = df_out_stock.rename(columns={final_stock_col: "final_stock"})
    df_out_stock["final_stock"] = _to_number_series(df_out_stock["final_stock"], compare_config=compare_config)

    if stok_col is not None and stok_col in df_out.columns:
        df_out_stock["output_stok"] = _to_number_series(df_out[stok_col], compare_config=compare_config)
    else:
        df_out_stock["output_stok"] = np.nan

    # Read top100
    if top_file.suffix.lower() in (".xlsx", ".xls"):
        df_top = pd.read_excel(top_file, dtype=str)
    else:
        df_top = pd.read_csv(top_file, dtype=str, keep_default_na=False)

    sku_col = _find_col(df_top, ["SKU", "Sku"])
    toko_col = _find_col(df_top, ["Toko", "Store", "Nama Store", "Outlet"])
    if sku_col is None:
        return pd.DataFrame([{ "error": f"Top100 file missing SKU column: {top_file.name}" }])

    has_toko = toko_col is not None

    df_top = df_top.rename(columns={sku_col: "SKU"})
    if has_toko:
        df_top = df_top.rename(columns={toko_col: "Toko"})
    else:
        df_top["Toko"] = ""

    df_top = _apply_key_normalization(df_top, compare_config=compare_config)

    running_date_str = datetime.strptime(date_str, "%Y%m%d").strftime("%d %b %Y")
    matched_date_cols = _find_date_header_columns(df_top, running_date_str)
    if not matched_date_cols:
        return pd.DataFrame([
            {
                "error": f"No date header matched running date '{running_date_str}' in top100 file {top_file.name}",
                "available_headers": ", ".join([str(c) for c in df_top.columns[:50]]),
            }
        ])

    date_col = matched_date_cols[0]
    top100_stock_col_name = f"top100_stock_{date_str}"

    df_top_stock = df_top[["SKU", "Toko", date_col]].copy()
    df_top_stock["top100_stock_raw"] = df_top_stock[date_col].astype(str).str.strip()
    df_top_stock["top100_stock_raw"] = df_top_stock["top100_stock_raw"].replace({"nan": "", "None": ""})

    df_top_stock[top100_stock_col_name] = _to_number_series(df_top_stock[date_col], compare_config=compare_config)
    df_top_stock = df_top_stock.drop(columns=[date_col])

    # Merge only keeping output top100 rows (left join)
    if has_toko:
        merged = df_out_stock.merge(df_top_stock, on=["SKU", "Toko"], how="left", indicator=True)
    else:
        merged = df_out_stock.merge(df_top_stock.drop(columns=["Toko"]), on=["SKU"], how="left", indicator=True)

    top_val = merged.get(top100_stock_col_name)
    top_blank = merged.get("top100_stock_raw").astype(str).str.strip().eq("") if "top100_stock_raw" in merged.columns else top_val.isna()

    expected = np.where(top_blank.to_numpy(), merged.get("output_stok").to_numpy(), top_val.to_numpy())
    expected = pd.Series(expected, index=merged.index)

    merged["comparison_basis"] = np.where(top_blank, "output_stok", "top100")
    merged["same_stock"] = (merged["final_stock"].fillna(0) - expected.fillna(0)).abs().le(1e-6)

    cols = [
        "SKU",
        "Toko",
        "final_stock",
        "output_stok",
        top100_stock_col_name,
        "comparison_basis",
        "same_stock",
        "_merge",
    ]

    for c in cols:
        if c not in merged.columns:
            merged[c] = None

    merged = merged[cols]
    return merged


def _build_grouped_validation_df(flags_df: pd.DataFrame) -> pd.DataFrame:
    # Expect columns: SKU, Toko, _merge + input__/output__/same__ columns
    params = [
        "is_top_100",
        "Stok",
        "Daily Sales",
        "Max. Daily Sales",
        "Lead Time",
        "Max. Lead Time",
        "Sedang PO",
    ]

    out = flags_df[["SKU", "Toko"]].copy()

    for p in params:
        # Check existence safely
        if f"input__{p}" in flags_df.columns:
            out[(p, "input")] = flags_df[f"input__{p}"]
        if f"output__{p}" in flags_df.columns:
            out[(p, "output")] = flags_df[f"output__{p}"]
        if f"same__{p}" in flags_df.columns:
            out[(p, "same")] = flags_df[f"same__{p}"]
    
    # Add extra singular columns safely
    extra_singles = ["emergency_po_cost", "final_updated_regular_po_cost"]
    for ex in extra_singles:
        if ex in flags_df.columns:
            # We map this to (ColumnName, "") so it appears next to others but without a submodule
            out[(ex, "")] = flags_df[ex]

    new_cols = []
    for col in out.columns:
        if isinstance(col, tuple) and len(col) == 2:
            new_cols.append(col)
        else:
            new_cols.append((str(col), ""))

    out.columns = pd.MultiIndex.from_tuples(new_cols)

    new_cols = []
    for col in out.columns:
        if isinstance(col, tuple) and len(col) == 2:
            if col[1] == "":
                new_cols.append(col[0])
            else:
                 new_cols.append(f"{col[0]}_{col[1]}")
        else:
            new_cols.append(str(col))

    out.columns = new_cols

    return out

def _write_validation_sheet_grouped(writer: pd.ExcelWriter, sheet_name: str, grouped_df: pd.DataFrame):
    # Write data starting from row 3 (after headers)
    grouped_df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, merge_cells=False, startrow=2)

    ws = writer.sheets[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    cols = list(grouped_df.columns)

    group_spans: dict[str, list[int]] = {}
    for idx, col in enumerate(cols, start=1):
        if '_' in col and not col.startswith("emergency_po") and not col.startswith("final_updated"): 
             # Heuristic: split by underscore unless it's one of our new columns which have underscores
             # better approach: check if it ends with _input, _output, _same
             if col.endswith("_input"):
                  grp = col[:-6]
             elif col.endswith("_output"):
                  grp = col[:-7]
             elif col.endswith("_same"):
                  grp = col[:-5]
             else:
                  grp = col
        else:
            grp = col
        group_spans.setdefault(grp, []).append(idx)

    for grp, positions in group_spans.items():
        if grp in ("SKU", "Toko") or len(positions) == 1:
            c = positions[0]
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
            ws.cell(row=1, column=c).value = grp
            continue

        if len(positions) <= 1:
            continue

        start = min(positions)
        end = max(positions)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        ws.cell(row=1, column=start).value = grp

        # Set subheaders in row 2
        for pos in positions:
            col_name = cols[pos - 1]
            if col_name.endswith("_input"):
                sub = "input"
            elif col_name.endswith("_output"):
                sub = "output"
            elif col_name.endswith("_same"):
                sub = "same"
            else:
                sub = ""
            ws.cell(row=2, column=pos).value = sub

    ws.auto_filter.ref = ws.cell(row=2, column=3).coordinate + ":" + ws.cell(row=max_row, column=max_col).coordinate

def _export_store_xlsx(
    result_xlsx_path: Path,
    flags_df: pd.DataFrame,
    mismatches_df: pd.DataFrame,
    metrics_df: pd.DataFrame,
    top100_cmp_df: pd.DataFrame,
):
    result_xlsx_path = Path(result_xlsx_path)
    result_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"[EXPORT] Starting export to {result_xlsx_path}")
    if flags_df.empty:
        print("[EXPORT] ERROR: flags_df is empty")
        return result_xlsx_path

    grouped_df = _build_grouped_validation_df(flags_df)
    
    try:
        with pd.ExcelWriter(result_xlsx_path, engine="openpyxl") as writer:
            _write_validation_sheet_grouped(writer, "validation", grouped_df)
            mismatches_df.to_excel(writer, index=False, sheet_name="mismatches")
            ws2 = writer.sheets["mismatches"]
            ws2.auto_filter.ref = ws2.cell(row=1, column=1).coordinate + ":" + ws2.cell(row=ws2.max_row, column=ws2.max_column).coordinate
            metrics_df.to_excel(writer, index=False, sheet_name="metrics")
            ws3 = writer.sheets["metrics"]
            ws3.auto_filter.ref = ws3.cell(row=1, column=1).coordinate + ":" + ws3.cell(row=ws3.max_row, column=ws3.max_column).coordinate
            
            # Apply formatting to metrics sheet
            # Column 1 = metric name, Column 2 = value
            for row in range(2, ws3.max_row + 1):
                metric_name = ws3.cell(row=row, column=1).value
                if not metric_name:
                    continue
                
                m = str(metric_name).lower()
                cell_val = ws3.cell(row=row, column=2)
                
                # Apply only if value is likely numeric (not bool "True"/"False" which Excel handles, or strings)
                # But here we just set format, Excel applies it if it's a number.
                
                if "cost" in m:
                    cell_val.number_format = '"Rp" #,##0'
                elif any(x in m for x in ["qty", "stock", "count", "rows", "diff"]):
                    cell_val.number_format = '#,##0'

            top100_cmp_df.to_excel(writer, index=False, sheet_name="top100_comparison")
            ws4 = writer.sheets["top100_comparison"]
            ws4.auto_filter.ref = ws4.cell(row=1, column=1).coordinate + ":" + ws4.cell(row=ws4.max_row, column=ws4.max_column).coordinate
            print(f"[EXPORT] Finished writing to {result_xlsx_path}")
    except Exception as e:
        print(f"[EXPORT] ERROR writing to {result_xlsx_path}: {e}")
    return result_xlsx_path

def run_validation_for_date(
    date_str: str,
    input_base_dir: Path = BASE_DIR / "data/input",
    output_base_dir: Path = BASE_DIR / "output",
    output_subdir: str = "complete/csv",
    output_exts: list[str] | None = None,
    compare_config: dict | None = None,
    export_xlsx_per_store: bool = True,
) -> pd.DataFrame:
    input_files = _list_input_files_for_date(date_str, input_base_dir=input_base_dir)

    results_dir = BASE_DIR / f"data/validation/{date_str}"
    results_dir.mkdir(parents=True, exist_ok=True)

    summary_rows = []

    export_futures = []
    export_results: dict[str, dict[str, str | None]] = {}

    max_workers = int(os.environ.get("VALIDATION_EXPORT_WORKERS", os.environ.get("SAVE_WORKERS", "2")))

    executor = ThreadPoolExecutor(max_workers=max_workers) if export_xlsx_per_store else None

    # Helper for robust sum
    def _sum_series(df, col):
        return _to_number_series(df[col] if col in df.columns else pd.Series(dtype=float), compare_config=compare_config).sum()

    try:
        for in_file in input_files:
            store_name = _get_store_name_from_filename(in_file.name)
            print(f"[VALIDATION][START] store={store_name} file={in_file.name}")

            out_file = _resolve_output_file(
                in_file,
                date_str,
                output_base_dir=output_base_dir,
                output_subdir=output_subdir,
                output_exts=output_exts,
            )

            if out_file is None:
                print(f"[VALIDATION][RESULT] store={store_name} file={in_file.name} status=missing_output")
                export_results[in_file.name] = {"export_status": "not_scheduled", "export_error": None}
                summary_rows.append(
                    {
                        "file": in_file.name,
                        "status": "missing_output",
                        "output_file": None,
                        "matched_rows": 0,
                        "input_rows": 0,
                        "mismatch_rows": 0,
                        "sum_initial_po_qty": 0,
                        "sum_emergency_po_qty": 0,
                        "sum_emergency_po_cost": 0,
                        "sum_final_updated_po_qty": 0,
                        "sum_final_updated_po_cost": 0,
                        "plain_sum_emergency_po_cost": 0,
                        "plain_sum_final_updated_po_cost": 0,
                        "result_xlsx": None,
                        "export_status": "not_scheduled",
                        "export_error": None,
                    }
                )
                continue

            try:
                merged_df, mismatches_df = validate_po_fields(in_file, out_file, compare_config=compare_config)
                flags_df = build_validation_flags(merged_df, compare_config=compare_config)
                
                # metrics_df contains the exact sums we want to check
                metrics_df = _build_metrics_df(in_file, out_file, compare_config=compare_config)
                
                # We can extract sums from metrics_df to ensure consistency
                # metrics_df has "metric" and "value" columns.
                def _get_metric_val(m_df, name):
                    v = m_df.loc[m_df["metric"] == name, "value"]
                    return v.iloc[0] if not v.empty else 0.0

                sum_final_updated_po_cost = _get_metric_val(metrics_df, "sum_final_updated_po_cost")
                sum_emergency_po_cost = _get_metric_val(metrics_df, "sum_emergency_po_cost")
                
                # Re-read df_out for other sums if needed, or trust metrics_df for cost
                df_out = pd.read_csv(out_file, sep=";", dtype=str, keep_default_na=False)
                
                sum_initial_po_qty = _sum_series(df_out, "initial_qty_po")
                sum_emergency_po_qty = _sum_series(df_out, "emergency_po_qty")
                sum_final_updated_po_qty = _sum_series(df_out, "final_updated_regular_po_qty")

                # "plain_sum" using the newly FIXED _plain_sum_series function
                plain_sum_emergency_po_cost = _plain_sum_series(df_out, "emergency_po_cost", compare_config=compare_config)
                plain_sum_final_updated_po_cost = _plain_sum_series(df_out, "final_updated_regular_po_cost", compare_config=compare_config)

                top100_cmp_df = _build_top100_comparison_df(in_file, out_file, date_str=date_str, compare_config=compare_config)

                matched_rows = int((merged_df["_merge"] == "both").sum())
                mismatch_rows = int(len(mismatches_df))

                result_xlsx_path = None
                export_status = "not_scheduled"
                export_error = None

                if export_xlsx_per_store:
                    result_xlsx_path = results_dir / f"{in_file.stem}.xlsx"
                    fut = executor.submit(
                        _export_store_xlsx,
                        result_xlsx_path,
                        flags_df,
                        mismatches_df,
                        metrics_df,
                        top100_cmp_df,
                    )
                    export_futures.append((in_file.name, store_name, fut, str(result_xlsx_path)))
                    export_status = "scheduled"
                    export_results[in_file.name] = {"export_status": export_status, "export_error": None}
                    pending = len(export_futures)
                    print(f"[VALIDATION][EXPORT] store={store_name} file={in_file.name} scheduled={result_xlsx_path} pending={pending}")

                summary_rows.append(
                    {
                        "file": in_file.name,
                        "status": "ok",
                        "output_file": str(out_file),
                        "matched_rows": matched_rows,
                        "flags_rows": int(len(flags_df)),
                        "mismatch_rows": mismatch_rows,
                        "sum_initial_po_qty": sum_initial_po_qty,
                        "sum_emergency_po_qty": sum_emergency_po_qty,
                        "sum_emergency_po_cost": sum_emergency_po_cost,
                        "sum_final_updated_po_qty": sum_final_updated_po_qty,
                        "sum_final_updated_po_cost": sum_final_updated_po_cost,
                        "plain_sum_emergency_po_cost": plain_sum_emergency_po_cost,
                        "plain_sum_final_updated_po_cost": plain_sum_final_updated_po_cost,
                        "result_xlsx": str(result_xlsx_path) if result_xlsx_path else None,
                        "export_status": export_status,
                        "export_error": export_error,
                    }
                )

                print(f"[VALIDATION][RESULT] store={store_name} file={in_file.name} status=ok matched_rows={matched_rows} mismatch_rows={mismatch_rows}")

            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"[VALIDATION][RESULT] store={store_name} file={in_file.name} status=error err={str(e)}")
                export_results[in_file.name] = {"export_status": "not_scheduled", "export_error": None}
                summary_rows.append(
                    {
                        "file": in_file.name,
                        "status": f"error: {str(e)[:200]}",
                        "output_file": str(out_file),
                        "matched_rows": 0,
                        "input_rows": 0,
                        "mismatch_rows": 0,
                        "sum_initial_po_qty": 0,
                        "sum_emergency_po_qty": 0,
                        "sum_emergency_po_cost": 0,
                        "sum_final_updated_po_qty": 0,
                        "sum_final_updated_po_cost": 0,
                        "plain_sum_emergency_po_cost": 0,
                        "plain_sum_final_updated_po_cost": 0,
                        "result_xlsx": None,
                        "export_status": "not_scheduled",
                        "export_error": None,
                    }
                )

        if executor is not None:
            remaining = len(export_futures)
            for file_name, store_name, fut, out_path in export_futures:
                try:
                    fut.result()
                    export_results[file_name] = {"export_status": "ok", "export_error": None}
                    print(f"[VALIDATION][EXPORT_DONE] store={store_name} file={file_name} path={out_path} pending={remaining-1}")
                except Exception as e:
                    export_results[file_name] = {"export_status": "error", "export_error": str(e)[:500]}
                    print(f"[VALIDATION][EXPORT_ERROR] store={store_name} file={file_name} err={str(e)} pending={remaining-1}")
                finally:
                    remaining -= 1

    finally:
        if executor is not None:
            executor.shutdown(wait=True)

    summary_df = pd.DataFrame(summary_rows)

    numerical_cols = ["matched_rows", "input_rows", "mismatch_rows", "sum_initial_po_qty", "sum_emergency_po_qty", "sum_emergency_po_cost", "sum_final_updated_po_qty", "sum_final_updated_po_cost", "plain_sum_emergency_po_cost", "plain_sum_final_updated_po_cost"]
    total_row = {"file": "TOTAL"}

    for col in numerical_cols:
        if col in summary_df.columns:
            total_row[col] = summary_df[col].sum()
    summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    if not summary_df.empty:
        if "export_status" not in summary_df.columns:
            summary_df["export_status"] = None
        if "export_error" not in summary_df.columns:
            summary_df["export_error"] = None

        summary_df["export_status"] = summary_df.apply(
            lambda r: (export_results.get(r["file"]) or {}).get("export_status", r.get("export_status")),
            axis=1,
        )
        summary_df["export_error"] = summary_df.apply(
            lambda r: (export_results.get(r["file"]) or {}).get("export_error", r.get("export_error")),
            axis=1,
        )

    index_path = results_dir / f"_{date_str}_result_index.xlsx"
    with pd.ExcelWriter(index_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="index")

        ws = writer.sheets["index"]
        ws.auto_filter.ref = ws.cell(row=1, column=1).coordinate + ":" + ws.cell(row=ws.max_row, column=ws.max_column).coordinate

        cost_cols = ["sum_emergency_po_cost", "sum_final_updated_po_cost", "plain_sum_emergency_po_cost", "plain_sum_final_updated_po_cost"]
        for col_idx, col_name in enumerate(summary_df.columns, 1):
            if col_name in cost_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):  # Skip header row
                    ws[f"{col_letter}{row}"].number_format = '"Rp" #,##0'

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 2



    print(f"Saved index XLSX: {index_path}")
    return summary_df


if __name__ == "__main__":
    date_str = "20251229"  # YYYYMMDD

    compare_config = {
        "mode": "loose",
        "number_locale": "auto",
        "abs_tol": 1e-6,
        "rel_tol": 0.0,
        "treat_blank_as_zero": True,
    }

    print(f"Running validation for date {date_str}...")
    summary_df = run_validation_for_date(
        date_str=date_str,
        input_base_dir=BASE_DIR / "data/input",
        output_base_dir=BASE_DIR / "output",
        output_subdir="complete/csv",
        output_exts=[".csv"],
        compare_config=compare_config,
        export_xlsx_per_store=True,
    )
    print("Done.")
